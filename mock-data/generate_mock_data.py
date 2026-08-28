"""
Mock data generator for TACS MVP project.
Reproduces the ORIGINAL 4-file structure (2 domains split across files, 3rd domain in one multi-sheet file)
so that Power Query logic built against this mock data transfers to real source files unchanged.

Run: python generate_mock_data.py
Output: ./out/*.xlsx  (원본과 동일한 파일명 패턴, 시트 구조)
"""
import random
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook

random.seed(42)  # 재현 가능한 목데이터 (재실행해도 같은 결과)

OUT_DIR = Path(__file__).parent / "out"
OUT_DIR.mkdir(exist_ok=True)
TODAY = date(2026, 8, 28)

DEPARTMENTS = ["NW운용1팀", "NW운용2팀", "IDC운용팀", "Cloud사업팀"]
PO_DEPTS = ["Cloud사업팀", "Platform사업팀", "Infra사업팀"]
BA_DEPTS = ["BA1", "BA2", "BA3"]

STD_SERVICES = [
    ("SVC01", "KT Cloud", "KTCLD"),
    ("SVC02", "Cloud DB", "KTDB"),
    ("SVC03", "IPTV", "KTIPTV"),
    ("SVC04", "GiGAeyes", "KTGE"),
    ("SVC05", "AI Platform", "KTAI"),
    ("SVC06", "BCP", "KTBCP"),
]
UNIT_SERVICES = {
    "KT Cloud": ["Cloud Compute", "Cloud Storage"],
    "Cloud DB": ["Cloud DB-Prod", "Cloud DB-Dev"],
    "IPTV": ["IPTV-Head", "IPTV-VOD"],
    "GiGAeyes": ["GiGAeyes-Core", "GiGAeyes-Edge"],
    "AI Platform": ["AI-Training", "AI-Serving"],
    "BCP": ["BCP-DR", "BCP-Backup"],
}
GRADE_POOL = ["Critical"] * 15 + ["High"] * 30 + ["Medium"] * 40 + ["Low"] * 15

LINUX_OS = [("Linux", "Rocky Linux", "9.4"), ("Linux", "Rocky Linux", "8.9"),
            ("Linux", "RHEL", "8.8"), ("Linux", "RHEL", "7.9"),
            ("Linux", "Ubuntu", "22.04"), ("Linux", "Ubuntu", "20.04")]
WINDOWS_OS = [("Windows", "Windows Server", "2022"), ("Windows", "Windows Server", "2019"),
              ("Windows", "Windows Server", "2012 R2")]
EOS_VERSIONS = {("RHEL", "7.9"), ("Windows Server", "2012 R2")}

EXCEPTION_REASONS = ["Legacy OS"] * 40 + ["Vendor 제한"] * 25 + ["서비스 영향도"] * 20 + ["업그레이드 예정"] * 15


def rand_date(start: date, end: date) -> date:
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, max(delta, 0)))


def pick_weighted(pool):
    return random.choice(pool)


# ---------------------------------------------------------------------------
# FACT_Host  (500건) — 모든 것의 기준키
# ---------------------------------------------------------------------------
hosts = []
name_seq = {prefix: 0 for _, _, prefix in STD_SERVICES}
ip_counter = 0

for i in range(1, 501):
    host_id = f"HOST{i:05d}"
    std_code, std_name, prefix = random.choice(STD_SERVICES)
    name_seq[prefix] += 1
    host_name = f"{prefix}{name_seq[prefix]:03d}"

    unit_name = random.choice(UNIT_SERVICES[std_name])
    grade = pick_weighted(GRADE_POOL)

    is_linux = random.random() < 0.70
    os_family, os_name, os_ver = random.choice(LINUX_OS if is_linux else WINDOWS_OS)

    is_vm = random.random() < 0.80
    platform = "VMware" if is_vm else "Physical"
    cpu_type = "Intel" if random.random() < 0.80 else "AMD"
    phys_cpu = random.choice([1, 2])
    phys_core = random.choice([8, 16, 24, 32])
    logi_core = phys_core * 2
    mem_mb = random.choice([16384, 32768, 65536, 131072])

    ip_counter += 1
    rep_ip = f"10.{20 + STD_SERVICES.index((std_code, std_name, prefix))}.{(ip_counter // 254) + 1}.{(ip_counter % 254) + 1}"

    op_dept = random.choice(DEPARTMENTS)
    po_dept = random.choice(PO_DEPTS)
    ba_dept = random.choice(BA_DEPTS)

    stale_days = random.choice([0, 0, 0, 1, 2, 3, 15, 30, 60]) if random.random() < 0.95 else random.randint(30, 90)
    last_seen = TODAY - timedelta(days=stale_days)

    facility_status = random.choices(["운영중", "점검중", "폐기예정"], weights=[90, 5, 5])[0]
    vm_collect = ("Agent" if random.random() < 0.9 else "Agentless") if is_vm else "-"

    hosts.append({
        "HostID": host_id,
        "호스트명": host_name,
        "대표IP": rep_ip,
        "대표IP노마스킹": rep_ip,
        "OS": os_name,
        "OS계열": os_family,
        "OS버전": os_ver,
        "Platform": platform,
        "CPU스펙": f"{cpu_type} Xeon/EPYC Series",
        "CPU유형": cpu_type,
        "물리적CPU수": phys_cpu,
        "CPU물리Core수": phys_core,
        "CPU논리Core수": logi_core,
        "메모리용량MB": mem_mb,
        "호스트운용부서": op_dept,
        "호스트운용자": f"운영자{random.randint(1, 30):02d}",
        "PO부서": po_dept,
        "PO": f"PO{random.randint(1, 15):02d}",
        "BA부서": ba_dept,
        "BA": f"BA사원{random.randint(1, 15):02d}",
        "표준서비스": std_code,
        "표준서비스명": std_name,
        "단위서비스": f"{std_code}-{UNIT_SERVICES[std_name].index(unit_name)+1}",
        "단위서비스명": unit_name,
        "서비스등급": grade,
        "VM수집방식상세": vm_collect,
        "설비상태": facility_status,
        "최근접속일자": last_seen,
        "미접속일수": stale_days,
    })

# ---------------------------------------------------------------------------
# DIM_Service / DIM_Department / DIM_OS — Host에서 distinct 추출
# ---------------------------------------------------------------------------
dim_service_rows = []
skey = 0
for std_code, std_name, _ in STD_SERVICES:
    for unit_name in UNIT_SERVICES[std_name]:
        skey += 1
        grades_used = sorted({h["서비스등급"] for h in hosts if h["표준서비스명"] == std_name and h["단위서비스명"] == unit_name}) or ["Medium"]
        dim_service_rows.append({
            "ServiceKey": f"SVCKEY{skey:03d}",
            "표준서비스": std_code, "표준서비스명": std_name,
            "단위서비스명": unit_name, "서비스등급대표값": grades_used[0],
        })

dim_dept_rows = []
dkey = 0
for d in sorted(set(DEPARTMENTS) | set(PO_DEPTS) | set(BA_DEPTS)):
    dkey += 1
    dim_dept_rows.append({"DepartmentKey": f"DEPTKEY{dkey:03d}", "부서명": d})

dim_os_rows = []
okey = 0
for fam, name, ver in sorted(set((h["OS계열"], h["OS"], h["OS버전"]) for h in hosts)):
    okey += 1
    dim_os_rows.append({
        "OSKey": f"OSKEY{okey:03d}", "OS계열": fam, "OS명": name, "버전": ver,
        "EOS대상여부": "Y" if (name, ver) in EOS_VERSIONS else "N",
    })

# ---------------------------------------------------------------------------
# FACT_Asset (100건) — 유선/무선 분리, 원본 파일 예시(KT-L4-SEOUL-01 등) 준수
# ---------------------------------------------------------------------------
REGIONS = ["SEOUL", "BUSAN", "DAEJEON", "GWANGJU", "INCHEON"]
WIRED_TYPES = ["Router", "Switch", "Firewall", "L4", "VPN"]
WIRELESS_TYPES = ["AP", "WLC", "Wireless-Bridge"]
TYPE_ABBR = {"Router": "RT", "Switch": "SW", "Firewall": "FW", "L4": "L4", "VPN": "VPN",
             "AP": "AP", "WLC": "WLC", "Wireless-Bridge": "WBR"}


def perturb_ip(ip: str) -> str:
    parts = ip.split(".")
    old_last = int(parts[-1])
    new_last = old_last
    while new_last == old_last:
        new_last = random.randint(1, 254)
    parts[-1] = str(new_last)
    return ".".join(parts)


def perturb_name(name: str) -> str:
    import re
    m = re.match(r"^([A-Za-z]+)(\d+)$", name)
    if not m:
        return name + "-X"
    letters, digits = m.groups()
    cut = max(len(letters) - 2, 1)
    return f"{letters[:cut]}-{letters[cut:]}{digits}"


all_asset_rows = []
for i in range(1, 101):
    wired = i <= 70  # 70 유선 / 30 무선
    dev_type = random.choice(WIRED_TYPES if wired else WIRELESS_TYPES)
    region = random.choice(REGIONS)
    seq = (i if wired else i - 70)
    device_name = f"KT-{TYPE_ABBR[dev_type]}-{region}-{seq:02d}"
    device_ip = f"10.10.{random.randint(1, 20)}.{random.randint(1, 254)}"
    is_linux = random.random() < 0.6
    os_family = "Linux" if is_linux else "Network OS"
    os_name, os_ver = ("Rocky Linux", "9.4") if is_linux else (random.choice(["IOS-XE", "JunOS", "PAN-OS"]), f"{random.randint(12,17)}.{random.randint(0,9)}")
    stale = random.choice([0, 1, 2, 5, 10]) if random.random() < 0.9 else random.randint(15, 45)

    row = {
        "AssetKey": f"ASSET{i:04d}",
        "SourceType": "유선" if wired else "무선",
        "장비그룹": random.choice(["네트워크부문", "보안부문", "인프라부문"]),
        "장비명": device_name,
        "장비IP": device_ip,
        "장비분류": "NETWORK",
        "장비유형": dev_type,
        "서비스구분": "운영" if random.random() < 0.85 else "개발",
        "자산관리자": random.choice(DEPARTMENTS),
        "OS계열": os_family,
        "OS": os_name,
        "버전": os_ver,
        "보안기능": "적용" if random.random() < 0.8 else "미적용",
        "동기화사용": "Y" if random.random() < 0.85 else "N",
        "동기화그룹명": f"NTP-GROUP-{random.randint(1,3):02d}",
        "장비설명": f"{region} IDC {dev_type}",
        "미접속일수": stale,
        "최근접속일자": TODAY - timedelta(days=stale),
        "등록방법": "자동" if random.random() < 0.9 else "수동",
        "등록일": rand_date(date(2024, 1, 1), date(2026, 6, 30)),
        "_wired": wired,
    }
    all_asset_rows.append(row)

# --- FACT_Host와의 겹침 주입: FACT_TACS_Matching이 100건 기준 80/5/5/10을 갖도록 ---
# (겹침 없이는 Power Query 매칭 로직이 항상 NO_MATCH만 내놓아 검증 대시보드가 무의미해짐)
overlap_hosts = random.sample(hosts, 90)
match_hosts, name_match_hosts, ip_match_hosts = overlap_hosts[:80], overlap_hosts[80:85], overlap_hosts[85:90]
random.shuffle(all_asset_rows)  # 어느 인덱스가 겹칠지 유선/무선에 고르게 섞이도록

for row, h in zip(all_asset_rows[0:80], match_hosts):
    row["장비명"], row["장비IP"] = h["호스트명"], h["대표IP"]
for row, h in zip(all_asset_rows[80:85], name_match_hosts):
    row["장비명"], row["장비IP"] = h["호스트명"], perturb_ip(h["대표IP"])
for row, h in zip(all_asset_rows[85:90], ip_match_hosts):
    row["장비명"], row["장비IP"] = perturb_name(h["호스트명"]), h["대표IP"]
# 나머지 all_asset_rows[90:100] 은 원래 생성된 독립 네트워크 장비명 그대로 → NO_MATCH

assets_wired = [{k: v for k, v in r.items() if k != "_wired"} for r in all_asset_rows if r["_wired"]]
assets_wireless = [{k: v for k, v in r.items() if k != "_wired"} for r in all_asset_rows if not r["_wired"]]

# ---------------------------------------------------------------------------
# FACT_TACS (500건, Host와 1:1) + FACT_SecurityCompliance (500건, Host와 1:1)
# 원본은 '지역NW운용본부' 한 파일 안에 시트로 분리되어 있음
# ---------------------------------------------------------------------------
tacs_rows, sec_rows = [], []
for idx, h in enumerate(hosts, start=1):
    os_accept = "Y" if random.random() < 0.90 else "N"
    db_accept = "Y" if random.random() < 0.85 else "N"
    os_reason = "" if os_accept == "Y" else pick_weighted(EXCEPTION_REASONS)
    db_reason = "" if db_accept == "Y" else pick_weighted(EXCEPTION_REASONS)

    tacs_rows.append({
        "HostID": h["HostID"], "호스트명": h["호스트명"], "장치ID": f"DEV{idx:05d}",
        "운영부서": h["호스트운용부서"], "표준서비스명": h["표준서비스명"], "단위서비스명": h["단위서비스명"],
        "TACS수용(OS)": os_accept, "TACS수용(DB)": db_accept,
        "TACS수용(OS)예외사유": os_reason, "TACS수용(DB)예외사유": db_reason,
        "운영상태": h["설비상태"], "설비상태": h["설비상태"],
        "VM여부": "Y" if h["Platform"] == "VMware" else "N",
    })

    is_eos = (h["OS"], h["OS버전"]) in EOS_VERSIONS
    webshell_agent = random.random() < 0.85
    av_installed = random.random() < 0.92
    edr_installed = random.random() < 0.78
    central_log = random.random() < 0.88
    smp_diag = random.random() < 0.80
    privacy_enc = random.random() < 0.75

    def yn(b):
        return "Y" if b else "N"

    sec_rows.append({
        "HostID": h["HostID"],
        "웹쉘Agent설치": yn(webshell_agent),
        "웹쉘정책설정": yn(webshell_agent and random.random() < 0.85),
        "서버백신설치": yn(av_installed),
        "서버백신실시간감시": yn(av_installed and random.random() < 0.90),
        "서버백신HIPS": yn(av_installed and random.random() < 0.70),
        "서버백신FullScan": yn(av_installed and random.random() < 0.80),
        "EDR설치": yn(edr_installed),
        "EDRFullScan": yn(edr_installed and random.random() < 0.75),
        "중앙로그관리수용": yn(central_log),
        "로그보관기간설정": random.choice(["90일", "180일", "365일"]) if central_log else "미설정",
        "ACL설정": yn(random.random() < 0.82),
        "EOS해소여부": yn(random.random() < (0.40 if is_eos else 0.95)),
        "SMP진단": yn(smp_diag),
        "SMP조치": yn(smp_diag and random.random() < 0.70),
        "개인정보암호화": yn(privacy_enc),
        "개인정보암호화일정": str(rand_date(date(2025, 1, 1), TODAY)) if privacy_enc else str(rand_date(TODAY, date(2027, 1, 1))),
    })


# ---------------------------------------------------------------------------
# 9장 확장: 서버관리_TACS등록여부 + TACS원본추출 (FACT_TACS_Verification 원천)
# 서버관리(FACT_Host)가 "TACS 등록됨"이라 주장하는 것과, TACS 시스템이 실제로
# 그 호스트를 (같은 이름/IP로) 갖고 있는지는 서로 다른 시스템의 서로 다른 데이터.
# UNVERIFIED는 TACS원본에 해당 HostID 행 자체가 없는 경우로 표현한다.
# ---------------------------------------------------------------------------
verify_order = list(hosts)
random.shuffle(verify_order)

N_NO_CLAIM = 50          # 서버관리_TACS등록여부 = N (전체 500건 중)
N_VERIFIED = 382         # 등록여부=Y(450건) 중 이름·IP 모두 일치
N_NAME_ONLY = 27         # 이름만 일치 (IP 다름)
N_IP_ONLY = 23           # IP만 일치 (이름 다름)
N_UNVERIFIED = 18        # 등록=Y라 주장하지만 TACS원본에 행 자체가 없음
assert N_VERIFIED + N_NAME_ONLY + N_IP_ONLY + N_UNVERIFIED == 500 - N_NO_CLAIM

claim_status = {}
cursor = 0
for h in verify_order[cursor:cursor + N_NO_CLAIM]:
    claim_status[h["HostID"]] = ("N", None)
cursor += N_NO_CLAIM
for h in verify_order[cursor:cursor + N_VERIFIED]:
    claim_status[h["HostID"]] = ("Y", "VERIFIED")
cursor += N_VERIFIED
for h in verify_order[cursor:cursor + N_NAME_ONLY]:
    claim_status[h["HostID"]] = ("Y", "VERIFIED_NAME_ONLY")
cursor += N_NAME_ONLY
for h in verify_order[cursor:cursor + N_IP_ONLY]:
    claim_status[h["HostID"]] = ("Y", "VERIFIED_IP_ONLY")
cursor += N_IP_ONLY
for h in verify_order[cursor:cursor + N_UNVERIFIED]:
    claim_status[h["HostID"]] = ("Y", "UNVERIFIED")

tacs_source_rows = []
for h in hosts:
    claim, plan_status = claim_status[h["HostID"]]
    h["서버관리_TACS등록여부"] = claim
    if plan_status in (None, "UNVERIFIED"):
        continue  # 등록 안 함 / 등록했다는 주장이 TACS원본엔 없음 -> 행 미생성
    if plan_status == "VERIFIED":
        src_name, src_ip = h["호스트명"], h["대표IP"]
    elif plan_status == "VERIFIED_NAME_ONLY":
        src_name, src_ip = h["호스트명"], perturb_ip(h["대표IP"])
    else:  # VERIFIED_IP_ONLY
        src_name, src_ip = perturb_name(h["호스트명"]), h["대표IP"]
    tacs_source_rows.append({
        "HostID": h["HostID"],
        "장비명": src_name,
        "IP": src_ip,
        "등록일시": str(rand_date(date(2024, 1, 1), TODAY)),
    })


# ---------------------------------------------------------------------------
# 10장 확장: TACS 관리 기준 확장
# 10-1) DIM_TACS_Scope_Rule (대상 판단 규칙, 12건)
# 10-2) FACT_TACS_Exception (예외 승인·만료, 200건)
# 10-3) FACT_TACS_History (이력/변경 스냅샷, ~2000건)
# 10-4) FACT_TACS_Escalation (미이행 리스크 에스컬레이션, 150건)
# ---------------------------------------------------------------------------

# --- 서비스구분(운영/개발)을 FACT_Host에 추가 (TACS대상여부 판단에 필요) ---
svcgu_order = list(hosts)
random.shuffle(svcgu_order)
n_dev = round(len(svcgu_order) * 0.15)
for h in svcgu_order[:n_dev]:
    h["서비스구분"] = "개발"
for h in svcgu_order[n_dev:]:
    h["서비스구분"] = "운영"

# --- 10-1) DIM_TACS_Scope_Rule ---
# 기본 원칙: 운영+Critical/High는 대상, 개발+Low는 비대상. Legacy OS는 별도 예외로 항상 비대상
# (FACT_Host[서비스등급]+[서비스구분] 조합 8건 + Legacy/기타 예외 사유 문서화용 4건 = 12건)
BASE_SCOPE_RULES = [
    ("Critical", "운영", "Y"), ("Critical", "개발", "Y"),
    ("High", "운영", "Y"), ("High", "개발", "N"),
    ("Medium", "운영", "Y"), ("Medium", "개발", "N"),
    ("Low", "운영", "N"), ("Low", "개발", "N"),
]
dim_scope_rule_rows = []
for i, (grade, gubun, target) in enumerate(BASE_SCOPE_RULES, start=1):
    dim_scope_rule_rows.append({
        "RuleKey": f"RULE{i:03d}",
        "적용조건_서비스등급": grade,
        "적용조건_OS계열": "ALL",
        "적용조건_서비스구분": gubun,
        "대상여부": target,
        "규칙설명": f"{gubun}환경 {grade}등급 기본 원칙",
        "최종개정일": rand_date(date(2025, 9, 1), TODAY),
    })
LEGACY_OVERRIDE_NOTES = [
    ("RHEL 7.9 (EOS)", "RHEL 7.9는 EOS 대상으로 TACS 적용 제외"),
    ("Windows2012 R2 (EOS)", "Windows2012 R2는 EOS 대상으로 TACS 적용 제외"),
    ("폐기예정 장비", "설비상태=폐기예정 장비는 TACS 적용 제외"),
    ("장기 미접속 장비", "미접속일수 90일 이상 장비는 TACS 적용 제외 검토 대상"),
]
for i, (os_label, desc) in enumerate(LEGACY_OVERRIDE_NOTES, start=len(BASE_SCOPE_RULES) + 1):
    dim_scope_rule_rows.append({
        "RuleKey": f"RULE{i:03d}",
        "적용조건_서비스등급": "ALL",
        "적용조건_OS계열": os_label,
        "적용조건_서비스구분": "ALL",
        "대상여부": "N",
        "규칙설명": desc,
        "최종개정일": rand_date(date(2025, 9, 1), TODAY),
    })

# --- 10-2) FACT_TACS_Exception (200건) ---
EXCEPTION_REASON_POOL = ["Legacy OS"] * 40 + ["Vendor 제한"] * 25 + ["서비스 영향도"] * 20 + ["업그레이드 예정"] * 15
REVIEW_MONTHS = {"Legacy OS": 12, "Vendor 제한": 6, "서비스 영향도": 6, "업그레이드 예정": 3}
APPROVERS = ["NW운용1팀장", "NW운용2팀장", "보안팀장", "IDC운용팀장"]
exception_host_pool = [t["HostID"] for t in tacs_rows if t["TACS수용(OS)"] == "N" or t["TACS수용(DB)"] == "N"]

exception_rows = []
for i in range(1, 201):
    host_id = random.choice(exception_host_pool)
    reason = pick_weighted(EXCEPTION_REASON_POOL)
    months = REVIEW_MONTHS[reason]
    review_days = months * 30
    r = random.random()
    if r < 0.20:
        # 이미 만료: 승인일을 review_days보다 더 오래 전으로 잡아 자연히 만료되게 함
        approve_date = TODAY - timedelta(days=review_days + random.randint(1, 200))
        expire_date = approve_date + timedelta(days=review_days)
    elif r < 0.35:
        # 만료임박(30일 이내): expire_date를 오늘 기준 0~30일 뒤로 역산
        expire_date = TODAY + timedelta(days=random.randint(0, 30))
        approve_date = expire_date - timedelta(days=review_days)
    else:
        # 유효: 승인일을 review_days 대비 충분히 최근으로 잡아 만료일이 30일 이상 남도록 함
        latest_approve_offset = max(review_days - 31, 0)
        approve_date = TODAY - timedelta(days=random.randint(0, latest_approve_offset))
        expire_date = approve_date + timedelta(days=review_days)
    exception_rows.append({
        "ExceptionKey": f"EXC{i:04d}",
        "HostID": host_id,
        "예외사유": reason,
        "예외사유상세": f"{reason} 관련 상세 사유",
        "승인자": random.choice(APPROVERS),
        "승인일": approve_date,
        "재검토주기_개월": months,
        "만료예정일": expire_date,
    })

# --- 10-3) FACT_TACS_History (약 2000건, 최근 90일 스냅샷 로그) ---
CHANGE_TYPE_POOL = (["신규등록"] * 15 + ["상태변경"] * 10 + ["예외추가"] * 20 +
                     ["예외만료"] * 15 + ["미확인전환"] * 10 + ["변경없음(정기 스냅샷)"] * 30)
VERIFY_STATUS_POOL_FOR_HISTORY = ["VERIFIED", "VERIFIED_NAME_ONLY", "VERIFIED_IP_ONLY"]

history_rows = []
for i in range(1, 2001):
    h = random.choice(hosts)
    change_type = pick_weighted(CHANGE_TYPE_POOL)
    snapshot_date = TODAY - timedelta(days=random.randint(0, 90))
    if change_type == "미확인전환":
        verify_status = "UNVERIFIED"
        prev_value = random.choice(VERIFY_STATUS_POOL_FOR_HISTORY)
    else:
        verify_status = random.choice(VERIFY_STATUS_POOL_FOR_HISTORY + ["UNVERIFIED"])
        prev_value = ""
    history_rows.append({
        "HistoryKey": f"HIST{i:05d}",
        "HostID": h["HostID"],
        "스냅샷일자": snapshot_date,
        "TACS수용_OS": "Y" if random.random() < 0.90 else "N",
        "TACS수용_DB": "Y" if random.random() < 0.85 else "N",
        "VerifyStatus": verify_status,
        "변경유형": change_type,
        "이전값": prev_value,
        "변경일": snapshot_date,
    })
history_rows.sort(key=lambda r: r["스냅샷일자"])

# --- 10-4) FACT_TACS_Escalation (150건) ---
RISK_SLA_DAYS = {"Critical": 3, "High": 7, "Medium": 14, "Low": 30}
RISK_POOL = ["Critical"] * 10 + ["High"] * 25 + ["Medium"] * 40 + ["Low"] * 25
unverified_host_ids = [hid for hid, (claim, status) in claim_status.items() if status == "UNVERIFIED"]
escalation_host_pool = (
    unverified_host_ids  # UNVERIFIED 18건
    + [e["HostID"] for e in exception_rows if e["만료예정일"] < TODAY]  # 만료된 예외건
)
if not escalation_host_pool:
    escalation_host_pool = [h["HostID"] for h in hosts]

escalation_rows = []
for i in range(1, 151):
    risk = pick_weighted(RISK_POOL)
    occur_date = TODAY - timedelta(days=random.randint(0, 60))
    sla_date = occur_date + timedelta(days=RISK_SLA_DAYS[risk])
    r = random.random()
    if r < 0.60:
        status = "완료"
    elif r < 0.85:
        status = "조치중"
    else:
        status = "미조치"
    final_action_date = None
    if status == "완료":
        offset = random.randint(-5, 5)
        final_action_date = sla_date + timedelta(days=offset)
    elif status == "미조치" and random.random() < 0.6:
        sla_date = TODAY - timedelta(days=random.randint(1, 10))  # SLA 초과 케이스 보장
    escalation_rows.append({
        "EscalationKey": f"ESC{i:04d}",
        "HostID": random.choice(escalation_host_pool),
        "리스크등급": risk,
        "발생일": occur_date,
        "SLA기한": sla_date,
        "담당자": f"담당자{random.randint(1, 20):02d}",
        "처리상태": status,
        "최종조치일": final_action_date,
    })


def write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])


def save_single(filename, sheet_name, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    write_sheet(ws, rows)
    wb.save(OUT_DIR / filename)


# 1) 장비 목록_유선 / 무선
save_single("장비 목록_유선_20260828.xlsx", "장비목록", assets_wired)
save_single("장비 목록_무선_20260828.xlsx", "장비목록", assets_wireless)

# 2) 호스트조회
save_single("호스트조회_20260828.xlsx", "호스트조회", hosts)

# 3) 지역NW운용본부 (멀티시트: TACS현황 + 보안통제)
wb = Workbook()
ws1 = wb.active
ws1.title = "서부"
write_sheet(ws1, tacs_rows)
ws2 = wb.create_sheet("보안통제")
write_sheet(ws2, sec_rows)
ws3 = wb.create_sheet("TACS원본추출")
write_sheet(ws3, tacs_source_rows)
wb.save(OUT_DIR / "지역NW운용본부_20260828.xlsx")

# 4) 차원 테이블 (Power BI에서 직접 로드할 참조 데이터 — 원본에 없는 파생 테이블)
wb = Workbook()
ws = wb.active
ws.title = "DIM_Service"
write_sheet(ws, dim_service_rows)
ws2 = wb.create_sheet("DIM_Department")
write_sheet(ws2, dim_dept_rows)
ws3 = wb.create_sheet("DIM_OS")
write_sheet(ws3, dim_os_rows)
ws4 = wb.create_sheet("DIM_TACS_Scope_Rule")
write_sheet(ws4, dim_scope_rule_rows)
wb.save(OUT_DIR / "차원테이블_20260828.xlsx")

# 5) TACS 관리대장 (10장 확장 — 예외/이력/에스컬레이션, 원본에 없는 관리 트래킹 데이터)
wb = Workbook()
ws1 = wb.active
ws1.title = "예외관리"
write_sheet(ws1, exception_rows)
ws2 = wb.create_sheet("이력관리")
write_sheet(ws2, history_rows)
ws3 = wb.create_sheet("에스컬레이션")
write_sheet(ws3, escalation_rows)
wb.save(OUT_DIR / "TACS_관리대장_20260828.xlsx")

# ---------------------------------------------------------------------------
# 매칭 검증 시뮬레이션 (Power Query 매칭 로직을 파이썬으로 미리 재현해 분포 확인)
# ---------------------------------------------------------------------------
host_by_name = {h["호스트명"]: h for h in hosts}
host_by_ip = {h["대표IP"]: h for h in hosts}
match_counts = {"MATCH": 0, "NAME_MATCH": 0, "IP_MATCH": 0, "NO_MATCH": 0}
for a in all_asset_rows:
    h_by_name = host_by_name.get(a["장비명"])
    h_by_ip = host_by_ip.get(a["장비IP"])
    if h_by_name and h_by_ip and h_by_name["HostID"] == h_by_ip["HostID"]:
        match_counts["MATCH"] += 1
    elif h_by_name:
        match_counts["NAME_MATCH"] += 1
    elif h_by_ip:
        match_counts["IP_MATCH"] += 1
    else:
        match_counts["NO_MATCH"] += 1

# --- FACT_TACS_Verification 시뮬레이션 (HostID로 TACS원본 조인, name/IP 비교) ---
tacs_src_by_host = {r["HostID"]: r for r in tacs_source_rows}
verify_counts = {"VERIFIED": 0, "VERIFIED_NAME_ONLY": 0, "VERIFIED_IP_ONLY": 0, "UNVERIFIED": 0, "해당없음(N)": 0}
for h in hosts:
    claim = h["서버관리_TACS등록여부"]
    src = tacs_src_by_host.get(h["HostID"])
    if src is None:
        verify_counts["UNVERIFIED" if claim == "Y" else "해당없음(N)"] += 1
    else:
        name_ok, ip_ok = src["장비명"] == h["호스트명"], src["IP"] == h["대표IP"]
        if name_ok and ip_ok:
            verify_counts["VERIFIED"] += 1
        elif name_ok:
            verify_counts["VERIFIED_NAME_ONLY"] += 1
        else:
            verify_counts["VERIFIED_IP_ONLY"] += 1

# ---------------------------------------------------------------------------
# 요약 리포트 (콘솔 한글 인코딩 문제 회피를 위해 UTF-8 파일로 기록)
# ---------------------------------------------------------------------------
with open(Path(__file__).parent / "summary.txt", "w", encoding="utf-8") as f:
    f.write(f"FACT_Host: {len(hosts)}건\n")
    f.write(f"  - Linux/Windows: {sum(1 for h in hosts if h['OS계열']=='Linux')}/{sum(1 for h in hosts if h['OS계열']=='Windows')}\n")
    f.write(f"  - VM/Physical: {sum(1 for h in hosts if h['Platform']=='VMware')}/{sum(1 for h in hosts if h['Platform']=='Physical')}\n")
    f.write(f"FACT_Asset: 유선 {len(assets_wired)}건, 무선 {len(assets_wireless)}건 (합계 {len(all_asset_rows)}건)\n")
    f.write(f"FACT_TACS: {len(tacs_rows)}건, OS수용Y {sum(1 for t in tacs_rows if t['TACS수용(OS)']=='Y')}건 ({sum(1 for t in tacs_rows if t['TACS수용(OS)']=='Y')/len(tacs_rows):.0%})\n")
    f.write(f"FACT_SecurityCompliance: {len(sec_rows)}건, EDR설치Y {sum(1 for s in sec_rows if s['EDR설치']=='Y')}건 ({sum(1 for s in sec_rows if s['EDR설치']=='Y')/len(sec_rows):.0%})\n")
    f.write(f"DIM_Service: {len(dim_service_rows)}건, DIM_Department: {len(dim_dept_rows)}건, DIM_OS: {len(dim_os_rows)}건\n")
    f.write(f"\nFACT_TACS_Matching 검증 시뮬레이션 (Asset 100건 기준):\n")
    for k, v in match_counts.items():
        f.write(f"  {k}: {v}건 ({v/len(all_asset_rows):.0%})\n")
    f.write(f"\nFACT_TACS_Verification 검증 시뮬레이션 (Host 500건 기준, TACS원본추출 {len(tacs_source_rows)}건):\n")
    for k, v in verify_counts.items():
        f.write(f"  {k}: {v}건 ({v/len(hosts):.0%})\n")

    f.write(f"\n--- 10장 확장 ---\n")
    f.write(f"FACT_Host 서비스구분: 운영 {sum(1 for h in hosts if h['서비스구분']=='운영')}건 / 개발 {sum(1 for h in hosts if h['서비스구분']=='개발')}건\n")
    f.write(f"DIM_TACS_Scope_Rule: {len(dim_scope_rule_rows)}건\n")
    f.write(f"FACT_TACS_Exception: {len(exception_rows)}건\n")
    n_expired = sum(1 for e in exception_rows if e['만료예정일'] < TODAY)
    n_soon = sum(1 for e in exception_rows if TODAY <= e['만료예정일'] <= TODAY + timedelta(days=30))
    f.write(f"  - 만료: {n_expired}건 ({n_expired/len(exception_rows):.0%}), 만료임박(30일 이내): {n_soon}건 ({n_soon/len(exception_rows):.0%})\n")
    f.write(f"FACT_TACS_History: {len(history_rows)}건 (최근 90일)\n")
    f.write(f"FACT_TACS_Escalation: {len(escalation_rows)}건\n")
    n_overdue = sum(1 for e in escalation_rows if e['처리상태'] != '완료' and e['SLA기한'] < TODAY)
    f.write(f"  - SLA 초과(미완료 & 기한경과): {n_overdue}건\n")

    f.write(f"\n출력 위치: {OUT_DIR}\n")

print("done - see summary.txt")
